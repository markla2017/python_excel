from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('files/Zeszyt1.xlsx')

ws = wb.active

# get values from excel cells
for row in range(1, 11):
    for col in range(1, 5):
        char = get_column_letter(col)
        print(ws[char + str(row)].value)

