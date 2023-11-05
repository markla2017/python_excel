from openpyxl import Workbook, load_workbook

fileName = 'files/Zeszyt1.xlsx'
wb = load_workbook(fileName) # You can open workbooks that are already open in Excel

ws = wb.active

print(wb.sheetnames)

# choose specific sheet
ws = wb['Arkusz2']

# create a new sheet
wb.create_sheet('Test') # requires to save in order to see changes in excel file
print(wb.sheetnames)

