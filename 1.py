from openpyxl import Workbook, load_workbook

fileName = 'files/Zeszyt1.xlsx'
wb = load_workbook(fileName) # You can open workbooks that are already open in Excel

ws = wb.active
print(ws) # name of active sheet
print(ws['A1'].value)

ws['A2'].value = 'Test' #changing value. This requires to save workbook

wb.save(fileName) # You can't SAVE a workbook if it is already open in excel.



