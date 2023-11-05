from openpyxl import Workbook, load_workbook

wb = load_workbook('files/tim.xlsx')
ws = wb.active

ws.merge_cells("A1:D1") # get data only from the first cells and remove values from other cells
ws.unmerge_cells("A1:D1")

wb.save('files/tim.xlsx')